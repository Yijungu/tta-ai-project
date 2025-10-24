from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.services.google_drive.templates import (  # noqa: E402
    SHARED_CRITERIA_NORMALIZED_NAMES,
    SPREADSHEET_RULES,
    build_default_shared_criteria_workbook,
    is_shared_criteria_candidate,
    normalize_shared_criteria_name,
)


def test_normalize_shared_criteria_name_strips_extension_and_whitespace() -> None:
    normalized = normalize_shared_criteria_name("  결함 판단 기준표 v1.0 .xlsx  ")
    assert normalized == normalize_shared_criteria_name("결함판단기준표v1.0")


def test_is_shared_criteria_candidate_matches_known_variants() -> None:
    for candidate in SHARED_CRITERIA_NORMALIZED_NAMES:
        assert is_shared_criteria_candidate(candidate + ".xlsx")
    assert not is_shared_criteria_candidate("임의파일.xlsx")


def test_build_default_shared_criteria_workbook_contains_required_headers() -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook  # pylint: disable=import-error

    workbook_bytes = build_default_shared_criteria_workbook()
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    workbook.close()
    assert headers == [
        "Invicti 결과",
        "결함 요약",
        "결함정도",
        "발생빈도",
        "품질특성",
        "결함 설명",
        "결함 제외 여부",
    ]


def test_spreadsheet_rules_configured_for_expected_menus() -> None:
    assert {"feature-list", "testcase-generation", "defect-report", "security-report"}.issubset(
        set(SPREADSHEET_RULES)
    )
    for rule in SPREADSHEET_RULES.values():
        assert "folder_name" in rule and "file_suffix" in rule and callable(rule["populate"])

