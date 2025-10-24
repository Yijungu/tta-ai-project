"""Feature list specific helpers."""

from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional, Sequence, Tuple

from fastapi import HTTPException

try:  # pragma: no cover - optional dependency guard
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

from ..excel_templates import (
    FEATURE_LIST_EXPECTED_HEADERS,
    extract_feature_list_overview,
    match_feature_list_header,
    populate_feature_list,
)
from .naming import drive_name_matches, looks_like_header_row
from .templates import FEATURE_LIST_SHEET_CANDIDATES, FEATURE_LIST_START_ROW


def parse_feature_list_workbook(workbook_bytes: bytes) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    if load_workbook is None:  # pragma: no cover - optional dependency guard
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")

    buffer = io.BytesIO(workbook_bytes)
    try:
        workbook = load_workbook(buffer, data_only=True)
    except Exception as exc:  # pragma: no cover - safety net
        raise HTTPException(status_code=500, detail="엑셀 파일을 읽는 중 오류가 발생했습니다.") from exc

    headers = list(FEATURE_LIST_EXPECTED_HEADERS)
    extracted_rows: List[Dict[str, str]] = []
    sheet_title = ""
    start_row = FEATURE_LIST_START_ROW
    header_row_values: Optional[Sequence[Any]] = None
    column_map: Dict[str, int] = {}

    try:
        sheet = workbook.active
        selected_title = sheet.title
        for candidate in FEATURE_LIST_SHEET_CANDIDATES:
            matched = False
            for title in workbook.sheetnames:
                if drive_name_matches(title, candidate):
                    try:
                        sheet = workbook[title]
                        selected_title = sheet.title
                        matched = True
                        break
                    except KeyError:
                        continue
            if matched:
                break

        sheet_title = selected_title

        max_col = sheet.max_column
        header_row_index: Optional[int] = None
        first_data_row_index: Optional[int] = None
        for idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_col=max_col, values_only=True),
            start=1,
        ):
            row_values: Sequence[Any] = row if isinstance(row, Sequence) else tuple()

            has_values = False
            for col_idx in range(len(headers)):
                cell_value = row_values[col_idx] if col_idx < len(row_values) else None
                if cell_value is None:
                    continue
                if str(cell_value).strip():
                    has_values = True
                    break

            header_match = looks_like_header_row(row_values, headers)

            if has_values and not header_match and first_data_row_index is None:
                first_data_row_index = idx

            if header_match:
                header_row_index = idx
                header_row_values = row_values
                break

            if idx >= FEATURE_LIST_START_ROW * 2 and first_data_row_index is not None:
                break

        if header_row_index is not None:
            start_row = header_row_index + 1
        elif first_data_row_index is not None:
            start_row = max(1, first_data_row_index)

        if header_row_values:
            display_headers = list(headers)
            for idx, value in enumerate(header_row_values):
                if value is None:
                    continue
                matched = match_feature_list_header(str(value))
                if matched and matched not in column_map:
                    column_map[matched] = idx
                    try:
                        header_index = headers.index(matched)
                    except ValueError:
                        header_index = None
                    if header_index is not None:
                        display_headers[header_index] = str(value).strip()

            headers = display_headers

        for default_idx, name in enumerate(FEATURE_LIST_EXPECTED_HEADERS):
            column_map.setdefault(name, default_idx)

        for row in sheet.iter_rows(
            min_row=max(1, start_row),
            max_col=max_col,
            values_only=True,
        ):
            row_values: Sequence[Any] = row if isinstance(row, Sequence) else tuple()

            if looks_like_header_row(row_values, headers):
                continue

            row_data: Dict[str, str] = {}
            has_values = False
            for header_name in headers:
                column_index = column_map.get(header_name)
                cell_value = (
                    row_values[column_index]
                    if column_index is not None and column_index < len(row_values)
                    else None
                )
                text = "" if cell_value is None else str(cell_value).strip()
                if text:
                    has_values = True
                row_data[header_name] = text

            if not has_values:
                continue

            description = row_data.get("기능 설명", "") or row_data.get("기능 개요", "")

            extracted_rows.append(
                {
                    "majorCategory": row_data.get("대분류", ""),
                    "middleCategory": row_data.get("중분류", ""),
                    "minorCategory": row_data.get("소분류", ""),
                    "featureDescription": description or "",
                }
            )
    finally:
        workbook.close()

    if not sheet_title:
        sheet_title = "기능리스트"

    overview_title, project_overview = extract_feature_list_overview(workbook_bytes)

    context = {
        "sheetName": sheet_title,
        "startRow": start_row,
        "headers": headers,
        "projectOverview": project_overview,
        "overviewSheetName": overview_title,
    }

    return context, extracted_rows


def build_feature_list_csv(rows: Sequence[Dict[str, str]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=list(FEATURE_LIST_EXPECTED_HEADERS),
        lineterminator="\n",
    )
    writer.writeheader()

    has_overview_column = "기능 개요" in FEATURE_LIST_EXPECTED_HEADERS

    for row in rows:
        major = str(row.get("majorCategory", "") or "").strip()
        middle = str(row.get("middleCategory", "") or "").strip()
        minor = str(row.get("minorCategory", "") or "").strip()
        description = str(row.get("featureDescription", "") or "").strip()

        if not any([major, middle, minor, description]):
            continue

        entry = {
            "대분류": major,
            "중분류": middle,
            "소분류": minor,
            "기능 설명": description,
        }
        if has_overview_column:
            entry["기능 개요"] = ""

        writer.writerow(entry)

    return output.getvalue()


def populate_workbook(
    workbook_bytes: bytes, csv_text: str, project_overview: Optional[str] = None
) -> bytes:
    return populate_feature_list(workbook_bytes, csv_text, project_overview)

